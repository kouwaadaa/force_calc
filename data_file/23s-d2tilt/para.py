import openpyxl
import math

def para_calc(file_name,num,tilt,v,alpha,clm):

    iteration = []
    data_num = []
    with open(file_name) as f:
        next(f)
        next(f)
        next(f)

        for line in f:
            data = line.split(' ')
            iteration.append(int(data[0]))
            data_num.append(float(data[1]))

    ave = sum(data_num[-num:])/num
    #ave = sum(data_num[801:900])/num

    #write excel
    excel_file = '23s-d2tilt.xlsx'

    book = openpyxl.load_workbook(excel_file)
    sheet = book['tilt' + str(tilt)]
    sheet.cell(row=6*v/5 + alpha/3 + 2,column=clm).value = ave                 
    book.save(excel_file)

    return ave

def renew(tilt,v,alpha):
    excel_file = '23s-d2tilt.xlsx'

    book = openpyxl.load_workbook(excel_file)
    sheet = book['tilt' + str(tilt)]
    lift_old = sheet.cell(row=6*v/5 + alpha/3 + 2,column=3).value      
    drag_old = sheet.cell(row=6*v/5 + alpha/3 + 2,column=4).value  

    sheet.cell(row=6*v/5 + alpha/3 + 2,column=3).value = -math.sin(math.radians(alpha))*drag_old + math.cos(math.radians(alpha))*lift_old
    sheet.cell(row=6*v/5 + alpha/3 + 2,column=4).value =  math.cos(math.radians(alpha))*drag_old + math.sin(math.radians(alpha))*lift_old 
      
    book.save(excel_file)
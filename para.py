import openpyxl

def para_calc(file_name,force,num,wing,alpha,clm):

    iteration = []
    data_num = []
    with open(file_name) as f:
        next(f)
        next(f)
        next(f)

        for line in f:
            data = line.split(' ')
            iteration.append(int(data[0]))
            data_num.append(float(data[1]))

    ave = sum(data_num[-num:])/num
    #ave = sum(data_num[801:900])/num
    print('%s:%f' % (force,ave))

    #write excel
    excel_file = 'wing_force.xlsx'

    book = openpyxl.load_workbook(excel_file)
    sheet = book['alpha=' + str(alpha)]
    # if wing == 1.0:
    #     if force == 'Drag':
    #         sheet.cell(row=2,column=clm).value = ave
    #     if force == 'Lift':
    #         sheet.cell(row=3,column=clm).value = ave
    #     if force == 'Moment':
    #         sheet.cell(row=4,column=clm).value = ave
    # if wing == 1.5:
    #     if force == 'Drag':
    #         sheet.cell(row=8,column=clm).value = ave
    #     if force == 'Lift':
    #         sheet.cell(row=9,column=clm).value = ave
    #     if force == 'Moment':
    #         sheet.cell(row=10,column=clm).value = ave
    # if wing == 1.5:
    #     if force == 'Drag':
    #         sheet.cell(row=14,column=clm).value = ave
    #     if force == 'Lift':
    #         sheet.cell(row=15,column=clm).value = ave
    #     if force == 'Moment':
    #         sheet.cell(row=16,column=clm).value = ave 
    if force == 'Drag':
        sheet.cell(row=20,column=clm).value = ave
    if force == 'Lift':
        sheet.cell(row=21,column=clm).value = ave
    if force == 'Moment':
        sheet.cell(row=22,column=clm).value = ave                      
    book.save(excel_file)


    return ave

def para_calc1(file_name,force,num):

    iteration = []
    data_num = []
    with open(file_name) as f:
        next(f)
        next(f)
        next(f)

        for line in f:
            data = line.split(' ')
            iteration.append(int(data[0]))
            data_num.append(float(data[1]))

    ave = sum(data_num[-num:])/num
    print('%s:%f' % (force,ave))

    return ave